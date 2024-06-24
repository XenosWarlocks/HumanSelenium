import csv
import os
import json
from openpyxl import Workbook, load_workbook
from collections import defaultdict

# Function to combine rows
def combine_rows(rows):
    combined = rows[0].copy()  # Start with the first row as a base
    for row in rows[1:]:
        for i in range(len(row)):
            if row[i] != '-' and not combined[i]:
                combined[i] = row[i]
    return combined

# Function to read and combine CSV data
def read_csv(input_file, output_folder):
    combined_data = defaultdict(list)

    # Construct input file path
    input_path = os.path.join(output_folder, input_file)

    # Read the CSV file
    with open(input_path, mode='r', newline='', encoding='utf-8') as infile:
        reader = csv.reader(infile)
        header = next(reader)  # Read the header
        for row in reader:
            name = row[0]
            # Append each row to the corresponding name key
            combined_data[name].append([cell if cell != '-' else '' for cell in row])

    return combined_data, header

# Function to read and combine JSON data
def read_json(input_file, output_folder):
    combined_data = defaultdict(list)

    # Construct input file path
    input_path = os.path.join(output_folder, input_file)

    # Read the JSON file
    with open(input_path, 'r') as infile:
        data = json.load(infile)
        header = data[0]  # Assuming the first row contains the header
        for row in data[1:]:
            name = row[0]
            combined_data[name].append([cell if cell != '-' else '' for cell in row])

    return combined_data, header

# Function to read and combine Excel data
def read_excel(input_file, output_folder):
    combined_data = defaultdict(list)

    # Construct input file path
    input_path = os.path.join(output_folder, input_file)

    # Load the Excel workbook
    wb = load_workbook(input_path)
    ws = wb.active

    # Read the Excel data
    header = [cell.value for cell in ws[1]]  # Assuming header is in the first row
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        combined_data[name].append([cell if cell != '-' else '' for cell in row])

    return combined_data, header

# Function to combine and write data to CSV
def combine_to_csv(input_file, output_folder):
    combined_data, header = read_csv(input_file, output_folder)

    merged_rows = []
    for name, rows in combined_data.items():
        merged_row = combine_rows(rows)
        merged_rows.append(merged_row)

    # Write the combined data to a new CSV file in the output folder
    output_csv = 'output.csv'
    output_path = os.path.join(output_folder, output_csv)
    with open(output_path, mode='w', newline='', encoding='utf-8') as outfile:
        writer = csv.writer(outfile)
        writer.writerow(header)  # Write the header
        for row in merged_rows:
            writer.writerow(row)

    print(f"Output written to {output_path}")

# Function to combine and write data to JSON
def combine_to_json(input_file, output_folder):
    combined_data, header = read_json(input_file, output_folder)

    merged_rows = []
    for name, rows in combined_data.items():
        merged_row = combine_rows(rows)
        merged_rows.append(merged_row)

    # Write the combined data to a new JSON file in the output folder
    output_json = 'output.json'
    output_path = os.path.join(output_folder, output_json)
    with open(output_path, mode='w', newline='', encoding='utf-8') as outfile:
        json.dump([header] + merged_rows, outfile, indent=4)

    print(f"Output written to {output_path}")

# Function to combine and write data to Excel
def combine_to_excel(input_file, output_folder):
    combined_data, header = read_excel(input_file, output_folder)

    merged_rows = []
    for name, rows in combined_data.items():
        merged_row = combine_rows(rows)
        merged_rows.append(merged_row)

    # Write the combined data to a new Excel file in the output folder
    output_excel = 'output.xlsx'
    output_path = os.path.join(output_folder, output_excel)
    wb = Workbook()
    ws = wb.active
    ws.append(header)  # Write the header
    for row in merged_rows:
        ws.append(row)
    wb.save(output_path)

    print(f"Output written to {output_path}")

# Function to determine file type and call appropriate combination function
def combine_data(input_file, output_folder, file_type):
    file_extension = file_type.lower()

    if file_extension == 'csv':
        combine_to_csv(input_file, output_folder)
    elif file_extension == 'json':
        combine_to_json(input_file, output_folder)
    elif file_extension == 'excel':
        combine_to_excel(input_file, output_folder)
    else:
        print("Unsupported file type. Please provide 'csv', 'json', or 'excel'.")
        return

# Main script execution
if __name__ == "__main__":
    # Ask user for input file and file type
    input_file = input("Enter the input file name: ")
    file_type = input("Enter the file type (.csv, .json, .xlsx): ")

    # Specify the output folder
    output_folder = 'gen_files'  # Desired output folder name

    # Run the combination process
    combine_data(input_file, output_folder, file_type)
